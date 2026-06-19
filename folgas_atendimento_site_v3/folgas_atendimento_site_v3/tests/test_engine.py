from datetime import date
from escala_engine import load_csv, generate_schedule, covered_periods


def test_generate_schedule_runs():
    colaboradores = load_csv('colaboradores.csv')
    quadro = load_csv('quadro_ideal.csv')
    eventos = load_csv('eventos.csv')
    ajustes = load_csv('ajustes_semanais.csv')
    schedule, summary = generate_schedule(colaboradores, quadro, eventos, ajustes, date(2026, 7, 4))
    assert not schedule.empty
    assert not summary.empty
    assert 'nome' in schedule.columns
    assert 'faltam' in summary.columns


def test_visual_excel_export_has_required_sheets():
    from io import BytesIO
    from openpyxl import load_workbook
    from escala_engine import to_excel_bytes

    colaboradores = load_csv('colaboradores.csv')
    quadro = load_csv('quadro_ideal.csv')
    eventos = load_csv('eventos.csv')
    ajustes = load_csv('ajustes_semanais.csv')
    start = date(2026, 7, 4)
    schedule, summary = generate_schedule(colaboradores, quadro, eventos, ajustes, start)

    payload = to_excel_bytes(schedule, summary, colaboradores=colaboradores, eventos=eventos, start=start)
    workbook = load_workbook(BytesIO(payload))

    assert workbook.sheetnames == ['Escala Semanal Visual', 'Escala por Colaborador', 'Escala por Dia', 'Cobertura']
    assert workbook['Escala Semanal Visual']['A1'].value == 'SÁBADO 04/07'
    assert workbook['Escala Semanal Visual']['A2'].value.startswith('Meio Dia')
    assert workbook['Escala Semanal Visual']['D2'].value.startswith('Tarde')
    assert workbook['Escala Semanal Visual']['G2'].value.startswith('Noite')
    assert workbook['Escala por Colaborador'].freeze_panes == 'A2'
    assert [cell.value for cell in workbook['Cobertura'][1]] == ['Dia', 'Período', 'Setor', 'Ideal', 'Escalado', 'Diferença', 'Status']
    assert workbook['Escala por Dia'].max_row > 1


def test_workload_controls_covered_periods():
    import pandas as pd

    assert covered_periods(pd.Series({'periodo': 'Manhã', 'carga_horaria': '7'})) == ['Manhã', 'Tarde']
    assert covered_periods(pd.Series({'periodo': 'Tarde', 'carga_horaria': '7'})) == ['Tarde', 'Noite']
    assert covered_periods(pd.Series({'periodo': 'Tarde', 'carga_horaria': '4'})) == ['Tarde']
def test_auto_fill_uses_secondary_sector_before_alerting():
    import pandas as pd

    colaboradores = pd.DataFrame([
        {
            'nome': 'Pessoa Apoio',
            'tipo': 'Fixo',
            'setor_cadastro': 'Copa',
            'setor_escala': 'Copa',
            'funcao': 'Atendente',
            'setores_secundarios': 'Praça',
            'horario_padrao': '08:00',
            'periodo_padrao': 'Manhã',
            'folga_fixa': 'Não',
            'trabalha_domingo': 'Sim',
            'status': 'Ativo',
        }
    ])
    quadro = pd.DataFrame([{'dia': 'Sábado', 'periodo': 'Manhã', 'Praça': '1', 'Copa': '0', 'Caixa': '0', 'Escritório': '0', 'Entrega': '0'}])

    schedule, summary = generate_schedule(colaboradores, quadro, pd.DataFrame(), pd.DataFrame(), date(2026, 7, 4))

    auto_rows = schedule[schedule['origem'] == 'Auto - fixo outro setor']
    assert not auto_rows.empty
    assert auto_rows.iloc[0]['setor'] == 'Praça'
    assert int(summary['faltam'].sum()) == 0


def test_missing_alert_has_reason_after_all_options():
    import pandas as pd

    colaboradores = pd.DataFrame(columns=['nome', 'tipo', 'setor_cadastro', 'setor_escala', 'funcao', 'setores_secundarios', 'horario_padrao', 'periodo_padrao', 'folga_fixa', 'trabalha_domingo', 'status'])
    quadro = pd.DataFrame([{'dia': 'Sábado', 'periodo': 'Manhã', 'Praça': '1', 'Copa': '0', 'Caixa': '0', 'Escritório': '0', 'Entrega': '0'}])

    _schedule, summary = generate_schedule(colaboradores, quadro, pd.DataFrame(), pd.DataFrame(), date(2026, 7, 4))

    row = summary.iloc[0]
    assert row['dia'] == 'Sábado'
    assert row['periodo'] == 'Manhã'
    assert row['setor'] == 'Praça'
    assert row['ideal'] == 1
    assert row['escalado'] == 0
    assert row['faltam'] == 1
    assert row['motivo_falta']
