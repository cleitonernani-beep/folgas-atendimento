from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DATA_DIR = Path(__file__).parent / "data"
DIAS_PT = {
    0: "Segunda",
    1: "Terça",
    2: "Quarta",
    3: "Quinta",
    4: "Sexta",
    5: "Sábado",
    6: "Domingo",
}
PERIODOS = ["Manhã", "Tarde", "Noite"]
SETORES = ["Praça", "Copa", "Caixa", "Escritório", "Entrega"]


def load_csv(name: str) -> pd.DataFrame:
    path = DATA_DIR / name
    if not path.exists():
        return pd.DataFrame()
    return pd.read_csv(path, dtype=str).fillna("")


def save_csv(df: pd.DataFrame, name: str) -> None:
    DATA_DIR.mkdir(exist_ok=True)
    df.to_csv(DATA_DIR / name, index=False)


def normalize_bool(value: object) -> bool:
    text = str(value).strip().lower()
    return text in {"sim", "s", "yes", "true", "1", "poderá trabalhar sim"}


def normalize_day(value: object) -> str:
    text = str(value).strip().lower()
    mapping = {
        "segunda": "Segunda",
        "segunda-feira": "Segunda",
        "terça": "Terça",
        "terca": "Terça",
        "terça-feira": "Terça",
        "terca-feira": "Terça",
        "quarta": "Quarta",
        "quarta-feira": "Quarta",
        "quinta": "Quinta",
        "quinta-feira": "Quinta",
        "sexta": "Sexta",
        "sexta-feira": "Sexta",
        "sábado": "Sábado",
        "sabado": "Sábado",
        "domingo": "Domingo",
        "não": "Não",
        "nao": "Não",
        "": "",
    }
    return mapping.get(text, str(value).strip())


def parse_date(value: object) -> date | None:
    text = str(value).strip()
    if not text:
        return None
    try:
        # A interface usa formato brasileiro (dd/mm/aaaa), mas mantemos
        # compatibilidade com datas ISO já existentes nos CSVs.
        return pd.to_datetime(text, dayfirst=True).date()
    except Exception:
        return None


def date_range(start: date, days: int = 7) -> list[date]:
    return [start + timedelta(days=i) for i in range(days)]


def is_between(current: date, start: date | None, end: date | None) -> bool:
    if start is None:
        return False
    if end is None:
        end = start
    return start <= current <= end


def employee_active(row: pd.Series, current: date, eventos: pd.DataFrame) -> bool:
    status = str(row.get("status", "Ativo")).strip().lower()
    if status and status not in {"ativo", "active"}:
        return False

    adm = parse_date(row.get("data_admissao", ""))
    dem = parse_date(row.get("data_desligamento", ""))
    if adm and current < adm:
        return False
    if dem and current > dem:
        return False

    nome = row.get("nome", "")
    if not eventos.empty:
        evs = eventos[eventos["nome"].astype(str).str.strip().str.lower() == str(nome).strip().lower()]
        for _, ev in evs.iterrows():
            tipo = str(ev.get("tipo", "")).strip().upper()
            if tipo in {"FÉRIAS", "FERIAS", "AFASTAMENTO", "ATESTADO", "DESLIGAMENTO"}:
                if is_between(current, parse_date(ev.get("data_inicio")), parse_date(ev.get("data_fim"))):
                    return False
    return True


def employee_event_status(nome: str, current: date, eventos: pd.DataFrame) -> tuple[str, str]:
    """Retorna status visual de ausência do colaborador para o Excel."""
    if eventos.empty:
        return "", ""
    evs = eventos[eventos["nome"].astype(str).str.strip().str.lower() == str(nome).strip().lower()]
    for _, ev in evs.iterrows():
        tipo = str(ev.get("tipo", "")).strip().upper()
        if not is_between(current, parse_date(ev.get("data_inicio")), parse_date(ev.get("data_fim"))):
            continue
        observacao = str(ev.get("observacao", "")).strip()
        if tipo in {"FÉRIAS", "FERIAS"}:
            return "FÉRIAS", observacao
        if tipo in {"AFASTAMENTO", "ATESTADO"}:
            return "AFASTAMENTO", observacao
    return "", ""

def explicit_folga(nome: str, current: date, eventos: pd.DataFrame) -> bool:
    if eventos.empty:
        return False
    evs = eventos[eventos["nome"].astype(str).str.strip().str.lower() == str(nome).strip().lower()]
    for _, ev in evs.iterrows():
        tipo = str(ev.get("tipo", "")).strip().upper()
        start = parse_date(ev.get("data_inicio"))
        end = parse_date(ev.get("data_fim")) or start
        if tipo in {"FOLGA", "FOLGA SEMANAL", "SEM"} and is_between(current, start, end):
            return True
        if tipo in {"DOM", "DOM/SEM", "DIA/DOM/SEM"} and current.weekday() == 6 and is_between(current, start, end):
            return True
        # Folga da semana vinculada a DOM/SEM ou DIA/DOM/SEM
        dia_sem = normalize_day(ev.get("dia_folga_semana", ""))
        if tipo in {"DOM/SEM", "DIA/DOM/SEM"} and dia_sem and DIAS_PT[current.weekday()] == dia_sem:
            return True
    return False


def works_only_day(nome: str, current: date, eventos: pd.DataFrame) -> bool:
    """Regra DIA/DOM/SEM: trabalha só Manhã/Tarde na data indicada, sem Noite."""
    if eventos.empty:
        return False
    evs = eventos[eventos["nome"].astype(str).str.strip().str.lower() == str(nome).strip().lower()]
    for _, ev in evs.iterrows():
        tipo = str(ev.get("tipo", "")).strip().upper()
        if tipo == "DIA/DOM/SEM" and is_between(current, parse_date(ev.get("data_inicio")), parse_date(ev.get("data_inicio"))):
            return True
    return False


def manual_actions_for_date(ajustes: pd.DataFrame, current: date) -> pd.DataFrame:
    if ajustes.empty or "data" not in ajustes.columns:
        return pd.DataFrame(columns=ajustes.columns if not ajustes.empty else [])
    parsed = ajustes["data"].map(parse_date)
    return ajustes[parsed == current].copy()


def has_manual_escalar(nome: str, current: date, ajustes: pd.DataFrame) -> bool:
    acts = manual_actions_for_date(ajustes, current)
    if acts.empty:
        return False
    mask = acts["nome"].astype(str).str.strip().str.lower() == str(nome).strip().lower()
    mask &= acts["acao"].astype(str).str.strip().str.upper() == "ESCALAR"
    return bool(mask.any())


def default_sunday_time(row: pd.Series) -> tuple[str, str]:
    nome = str(row.get("nome", "")).strip().upper()
    horario = str(row.get("horario_padrao", "")).strip()
    periodo = str(row.get("periodo_padrao", "")).strip() or "Tarde"

    if nome in {"ANA JULYA", "THIAGO"}:
        return "18:00", "Noite"
    if horario < "16:00":
        return "16:00", "Tarde"
    return horario, periodo


def default_saturday_override(row: pd.Series) -> tuple[str, str] | None:
    nome = str(row.get("nome", "")).strip().upper()
    if nome == "LOHRAN":
        return "17:00", "Noite"
    return None


def should_schedule_employee(row: pd.Series, current: date, eventos: pd.DataFrame, ajustes: pd.DataFrame, domingo_especial: bool) -> bool:
    nome = str(row.get("nome", "")).strip()
    tipo = str(row.get("tipo", "")).strip().lower()
    dia = DIAS_PT[current.weekday()]

    if not employee_active(row, current, eventos):
        return False

    # Extras entram por sugestão para fechar quadro, não como base fixa.
    if tipo == "extra":
        return False

    # Domingo: estagiários não entram e folga fixa domingo é padrão, mas pode ser forçada por ajuste manual.
    if current.weekday() == 6:
        if tipo.startswith("estagi"):
            return False
        if not normalize_bool(row.get("trabalha_domingo", "")):
            return False
        if normalize_day(row.get("folga_fixa", "")) == "Domingo" and not has_manual_escalar(nome, current, ajustes):
            return False

    if explicit_folga(nome, current, eventos) and not has_manual_escalar(nome, current, ajustes):
        return False

    folga_fixa = normalize_day(row.get("folga_fixa", ""))
    if folga_fixa and folga_fixa != "Não" and dia == folga_fixa and not has_manual_escalar(nome, current, ajustes):
        return False

    # Domingo normal não tem Manhã.
    if current.weekday() == 6 and not domingo_especial:
        return True
    return True


def generate_base_schedule(
    colaboradores: pd.DataFrame,
    eventos: pd.DataFrame,
    ajustes: pd.DataFrame,
    start: date,
    domingo_especial: bool = False,
) -> pd.DataFrame:
    rows: list[dict] = []
    for current in date_range(start, 7):
        dia = DIAS_PT[current.weekday()]
        for _, emp in colaboradores.iterrows():
            if not should_schedule_employee(emp, current, eventos, ajustes, domingo_especial):
                continue
            nome = str(emp.get("nome", "")).strip()
            horario = str(emp.get("horario_padrao", "")).strip()
            periodo = str(emp.get("periodo_padrao", "")).strip() or "Tarde"
            setor = str(emp.get("setor_escala", emp.get("setor_cadastro", ""))).strip()
            funcao = str(emp.get("funcao", "")).strip()

            sat_override = default_saturday_override(emp) if current.weekday() == 5 else None
            if sat_override:
                horario, periodo = sat_override
            if current.weekday() == 6:
                horario, periodo = default_sunday_time(emp)
                if not domingo_especial and periodo == "Manhã":
                    periodo = "Tarde"
                    horario = "16:00"

            # DIA/DOM/SEM na data do DIA: não trabalha à noite.
            if works_only_day(nome, current, eventos) and periodo == "Noite":
                continue

            rows.append(
                {
                    "data": current.isoformat(),
                    "dia": dia,
                    "periodo": periodo,
                    "setor": setor,
                    "funcao": funcao,
                    "nome": nome,
                    "horario": horario,
                    "origem": "Base",
                    "observacao": str(emp.get("obs", "")).strip(),
                }
            )
    schedule = pd.DataFrame(rows)
    if schedule.empty:
        return pd.DataFrame(columns=["data", "dia", "periodo", "setor", "funcao", "nome", "horario", "origem", "observacao"])
    schedule = apply_manual_adjustments(schedule, colaboradores, ajustes)
    return schedule.sort_values(["data", "periodo", "setor", "horario", "nome"]).reset_index(drop=True)


def apply_manual_adjustments(schedule: pd.DataFrame, colaboradores: pd.DataFrame, ajustes: pd.DataFrame) -> pd.DataFrame:
    if ajustes.empty:
        return schedule
    schedule = schedule.copy()
    for _, act in ajustes.iterrows():
        current_date = parse_date(act.get("data", ""))
        current = current_date.isoformat() if current_date else str(act.get("data", "")).strip()
        nome = str(act.get("nome", "")).strip()
        acao = str(act.get("acao", "")).strip().upper()
        if not current or not nome or not acao:
            continue
        mask = (schedule["data"].astype(str) == current) & (schedule["nome"].astype(str).str.strip().str.lower() == nome.lower())
        if acao == "FOLGAR":
            schedule = schedule.loc[~mask].copy()
        elif acao in {"ALTERAR_HORARIO", "ALTERAR_SETOR", "ALTERAR_PERIODO"}:
            if mask.any():
                if str(act.get("setor", "")).strip():
                    schedule.loc[mask, "setor"] = str(act.get("setor", "")).strip()
                if str(act.get("periodo", "")).strip():
                    schedule.loc[mask, "periodo"] = str(act.get("periodo", "")).strip()
                if str(act.get("horario", "")).strip():
                    schedule.loc[mask, "horario"] = str(act.get("horario", "")).strip()
                if str(act.get("observacao", "")).strip():
                    schedule.loc[mask, "observacao"] = str(act.get("observacao", "")).strip()
        elif acao == "ESCALAR":
            emp = colaboradores[colaboradores["nome"].astype(str).str.strip().str.lower() == nome.lower()]
            funcao = ""
            if not emp.empty:
                funcao = str(emp.iloc[0].get("funcao", "")).strip()
            new_row = {
                "data": current,
                "dia": DIAS_PT[pd.to_datetime(current).weekday()],
                "periodo": str(act.get("periodo", "")).strip() or "Tarde",
                "setor": str(act.get("setor", "")).strip() or (str(emp.iloc[0].get("setor_escala", "")).strip() if not emp.empty else ""),
                "funcao": funcao,
                "nome": nome,
                "horario": str(act.get("horario", "")).strip() or (str(emp.iloc[0].get("horario_padrao", "")).strip() if not emp.empty else ""),
                "origem": "Manual",
                "observacao": str(act.get("observacao", "")).strip(),
            }
            # Evita duplicar se já houver linha idêntica de ajuste manual.
            duplicate = (
                (schedule["data"].astype(str) == new_row["data"])
                & (schedule["nome"].astype(str).str.lower() == new_row["nome"].lower())
                & (schedule["periodo"].astype(str) == new_row["periodo"])
            )
            if not duplicate.any():
                schedule = pd.concat([schedule, pd.DataFrame([new_row])], ignore_index=True)
    return schedule


def covered_periods(row: pd.Series) -> list[str]:
    """
    Calcula presença operacional por período sem usar horário de saída.

    Decisão do protótipo:
    - Entrada de Manhã conta para Manhã e Tarde.
    - Entrada de Tarde conta para Tarde e Noite.
    - Entrada de Noite conta para Noite.
    - Extras sugeridos de Manhã contam só Manhã, pois normalmente são apoio pontual.
    """
    periodo = str(row.get("periodo", "")).strip()
    origem = str(row.get("origem", "")).strip()
    if origem == "Sugestão extra" and periodo == "Manhã":
        return ["Manhã"]
    if periodo == "Manhã":
        return ["Manhã", "Tarde"]
    if periodo == "Tarde":
        return ["Tarde", "Noite"]
    if periodo == "Noite":
        return ["Noite"]
    return [periodo] if periodo else []


def coverage_frame(schedule: pd.DataFrame) -> pd.DataFrame:
    if schedule.empty:
        return pd.DataFrame(columns=["dia", "periodo", "setor", "nome"])
    rows = []
    for _, row in schedule.iterrows():
        for per in covered_periods(row):
            rows.append({
                "data": row.get("data", ""),
                "dia": row.get("dia", ""),
                "periodo": per,
                "setor": row.get("setor", ""),
                "nome": row.get("nome", ""),
            })
    return pd.DataFrame(rows)


def coverage_summary(schedule: pd.DataFrame, ideal: pd.DataFrame, reasons: dict[tuple[str, str, str], str] | None = None) -> pd.DataFrame:
    if ideal.empty:
        return pd.DataFrame()
    ideal = ideal.copy()
    ideal["dia"] = ideal["dia"].map(normalize_day)
    coverage = coverage_frame(schedule)
    rows = []
    reasons = reasons or {}
    for _, ir in ideal.iterrows():
        dia = ir.get("dia", "")
        periodo = ir.get("periodo", "")
        for setor in SETORES:
            try:
                needed = int(float(str(ir.get(setor, 0)).replace(",", ".") or 0))
            except Exception:
                needed = 0
            if needed <= 0:
                continue
            if coverage.empty:
                actual = 0
            else:
                actual = int(((coverage["dia"] == dia) & (coverage["periodo"] == periodo) & (coverage["setor"] == setor)).sum())
            missing = max(needed - actual, 0)
            rows.append({
                "dia": dia,
                "periodo": periodo,
                "setor": setor,
                "ideal": needed,
                "escalado": actual,
                "faltam": missing,
                "sobra": max(actual - needed, 0),
                "motivo_falta": reasons.get((dia, periodo, setor), "") if missing else "",
            })
    return pd.DataFrame(rows)


def normalized_sectors(value: object) -> set[str]:
    text = str(value).replace("/", ",").replace(";", ",")
    return {part.strip().lower() for part in text.split(",") if part.strip()}


def can_work_sector(row: pd.Series, setor: str, *, secondary: bool = False) -> bool:
    target = setor.strip().lower()
    own = str(row.get("setor_escala", row.get("setor_cadastro", ""))).strip().lower()
    if not secondary:
        return own == target
    return target in normalized_sectors(row.get("setores_secundarios", "")) and own != target


def is_internal_extra(row: pd.Series) -> bool:
    name = str(row.get("nome", "")).strip().lower()
    return name in {"daia - extra", "hemerson - extra"}


def target_time_and_period(row: pd.Series, current: date, periodo: str, domingo_especial: bool) -> tuple[str, str]:
    horario = str(row.get("horario_padrao", "")).strip()
    target_period = periodo
    if current.weekday() == 6 and periodo in {"Tarde", "Noite"}:
        horario, _ = default_sunday_time(row)
        if horario < "16:00" and not domingo_especial:
            horario = "16:00"
    sat_override = default_saturday_override(row) if current.weekday() == 5 else None
    if sat_override and periodo == sat_override[1]:
        horario, target_period = sat_override
    return horario, target_period



def covering_row_indices(schedule: pd.DataFrame, nome: str, current: date, periodo: str) -> list[int]:
    indices: list[int] = []
    if schedule.empty:
        return indices
    day_rows = schedule[
        (schedule["data"].astype(str) == current.isoformat())
        & (schedule["nome"].astype(str).str.strip().str.lower() == nome.strip().lower())
    ]
    for idx, row in day_rows.iterrows():
        if periodo in covered_periods(row):
            indices.append(idx)
    return indices


def can_reassign_covering_row(schedule: pd.DataFrame, ideal: pd.DataFrame, idx: int, target_setor: str, periodo: str) -> bool:
    row = schedule.loc[idx]
    current_setor = str(row.get("setor", "")).strip()
    dia = str(row.get("dia", "")).strip()
    if current_setor.lower() == target_setor.strip().lower():
        return False
    current_summary = coverage_summary(schedule, ideal)
    same = current_summary[
        (current_summary["dia"] == dia)
        & (current_summary["periodo"] == periodo)
        & (current_summary["setor"] == current_setor)
    ]
    if same.empty:
        return True
    return int(same.iloc[0].get("sobra", 0)) > 0 or int(same.iloc[0].get("ideal", 0)) == 0

def person_covers_period(schedule: pd.DataFrame, nome: str, current: date, periodo: str) -> bool:
    if schedule.empty:
        return False
    day_rows = schedule[
        (schedule["data"].astype(str) == current.isoformat())
        & (schedule["nome"].astype(str).str.strip().str.lower() == nome.strip().lower())
    ]
    return any(periodo in covered_periods(row) for _, row in day_rows.iterrows())


def can_be_auto_added(row: pd.Series, schedule: pd.DataFrame, current: date, periodo: str, eventos: pd.DataFrame, ajustes: pd.DataFrame, domingo_especial: bool) -> tuple[bool, str]:
    nome = str(row.get("nome", "")).strip()
    tipo = str(row.get("tipo", "")).strip().lower()
    dia = DIAS_PT[current.weekday()]

    if not employee_active(row, current, eventos):
        return False, "indisponível por status, férias, afastamento ou desligamento"
    if current.weekday() == 6 and tipo.startswith("estagi"):
        return False, "estagiários não trabalham domingo"
    if current.weekday() == 6 and not normalize_bool(row.get("trabalha_domingo", "")):
        return False, "não trabalha domingo"
    if current.weekday() == 6 and periodo == "Manhã" and not domingo_especial:
        return False, "domingo normal não tem manhã"
    if explicit_folga(nome, current, eventos) and not has_manual_escalar(nome, current, ajustes):
        return False, "folga/evento no dia"
    folga_fixa = normalize_day(row.get("folga_fixa", ""))
    if folga_fixa and folga_fixa != "Não" and dia == folga_fixa and not has_manual_escalar(nome, current, ajustes):
        return False, "folga fixa no dia"
    if works_only_day(nome, current, eventos) and periodo == "Noite":
        return False, "regra DIA/DOM/SEM não permite noite"
    if person_covers_period(schedule, nome, current, periodo):
        return False, "já cobre este dia/período"

    default_period = str(row.get("periodo_padrao", "")).strip() or "Tarde"
    # O sistema não antecipa horário sozinho: Manhã só entra automaticamente
    # para quem já tem Manhã como período padrão.
    if periodo == "Manhã" and default_period != "Manhã":
        return False, "cobrir Manhã exigiria antecipação manual"
    return True, ""


def add_schedule_row(schedule: pd.DataFrame, row: pd.Series, current: date, periodo: str, setor: str, origem: str, domingo_especial: bool) -> pd.DataFrame:
    horario, target_period = target_time_and_period(row, current, periodo, domingo_especial)
    addition = {
        "data": current.isoformat(),
        "dia": DIAS_PT[current.weekday()],
        "periodo": target_period,
        "setor": setor,
        "funcao": str(row.get("funcao", "Extra")).strip(),
        "nome": str(row.get("nome", "")).strip(),
        "horario": horario,
        "origem": origem,
        "observacao": f"Preenchido automaticamente para cobrir {setor} / {periodo}.",
    }
    return pd.concat([schedule, pd.DataFrame([addition])], ignore_index=True)


def auto_fill_coverage(
    schedule: pd.DataFrame,
    colaboradores: pd.DataFrame,
    ideal: pd.DataFrame,
    eventos: pd.DataFrame,
    ajustes: pd.DataFrame,
    start: date,
    domingo_especial: bool = False,
) -> tuple[pd.DataFrame, dict[tuple[str, str, str], str]]:
    """Preenche faltas antes de alertar, seguindo a ordem de prioridade do negócio."""
    schedule = schedule.copy()
    reasons: dict[tuple[str, str, str], str] = {}
    extra_count = {
        str(row.get("nome", "")).strip(): int((schedule["nome"].astype(str).str.strip().str.lower() == str(row.get("nome", "")).strip().lower()).sum())
        for _, row in colaboradores.iterrows()
        if str(row.get("tipo", "")).strip().lower() == "extra"
    }

    day_by_name = {DIAS_PT[current.weekday()]: current for current in date_range(start)}

    while True:
        summary = coverage_summary(schedule, ideal)
        gaps = summary[summary["faltam"] > 0] if not summary.empty else pd.DataFrame()
        if gaps.empty:
            break

        changed = False
        reasons.clear()
        for _, gap in gaps.iterrows():
            dia = str(gap["dia"])
            periodo = str(gap["periodo"])
            setor = str(gap["setor"])
            current = day_by_name.get(dia)
            if current is None:
                reasons[(dia, periodo, setor)] = "dia fora da semana operacional"
                continue
            if current.weekday() == 6 and periodo == "Manhã" and not domingo_especial:
                reasons[(dia, periodo, setor)] = "domingo normal não tem período da Manhã"
                continue

            attempts: list[str] = []
            stages = [
                ("Fixo próprio setor", colaboradores[colaboradores["tipo"].astype(str).str.strip().str.lower() != "extra"], lambda r: can_work_sector(r, setor, secondary=False), "Auto - fixo do setor"),
                ("Fixo setor secundário", colaboradores[colaboradores["tipo"].astype(str).str.strip().str.lower() != "extra"], lambda r: can_work_sector(r, setor, secondary=True), "Auto - fixo outro setor"),
                ("Extra interno", colaboradores[colaboradores.apply(is_internal_extra, axis=1)], lambda r: can_work_sector(r, setor, secondary=False) or can_work_sector(r, setor, secondary=True), "Auto - extra interno"),
                ("Extra externo", colaboradores[(colaboradores["tipo"].astype(str).str.strip().str.lower() == "extra") & (~colaboradores.apply(is_internal_extra, axis=1))], lambda r: can_work_sector(r, setor, secondary=False) or can_work_sector(r, setor, secondary=True), "Auto - extra externo"),
            ]

            filled = False
            for label, pool, sector_filter, origem in stages:
                candidates = []
                block_reasons: list[str] = []
                for _, candidate in pool.iterrows():
                    if not sector_filter(candidate):
                        continue
                    covering_indices = covering_row_indices(schedule, str(candidate.get("nome", "")).strip(), current, periodo)
                    reassign_indices = [idx for idx in covering_indices if can_reassign_covering_row(schedule, ideal, idx, setor, periodo)]
                    if reassign_indices:
                        candidate = candidate.copy()
                        candidate["_reassign_index"] = reassign_indices[0]
                        candidates.append(candidate)
                        continue
                    ok, reason = can_be_auto_added(candidate, schedule, current, periodo, eventos, ajustes, domingo_especial)
                    if ok:
                        candidate = candidate.copy()
                        candidate["_reassign_index"] = None
                        candidates.append(candidate)
                    elif reason:
                        block_reasons.append(reason)
                if not candidates:
                    attempts.append(f"{label}: sem candidato disponível" + (f" ({sorted(set(block_reasons))[0]})" if block_reasons else ""))
                    continue

                candidates_df = pd.DataFrame(candidates)
                if origem in {"Auto - extra interno", "Auto - extra externo"}:
                    candidates_df["_rodizio"] = candidates_df["nome"].astype(str).map(lambda n: extra_count.get(n.strip(), 0))
                    candidates_df = candidates_df.sort_values(["_rodizio", "nome"])
                else:
                    candidates_df = candidates_df.sort_values(["nome"])
                chosen = candidates_df.iloc[0]
                reassign_index = chosen.get("_reassign_index")
                if pd.notna(reassign_index):
                    schedule.loc[int(reassign_index), "setor"] = setor
                    schedule.loc[int(reassign_index), "origem"] = origem
                    schedule.loc[int(reassign_index), "observacao"] = f"Realocado automaticamente para cobrir {setor} / {periodo}."
                else:
                    schedule = add_schedule_row(schedule, chosen, current, periodo, setor, origem, domingo_especial)
                chosen_name = str(chosen.get("nome", "")).strip()
                if chosen_name in extra_count:
                    extra_count[chosen_name] += 1
                changed = True
                filled = True
                break

            if not filled:
                reasons[(dia, periodo, setor)] = "; ".join(attempts) or "nenhuma opção de preenchimento encontrada"

        if not changed:
            break

    return schedule.sort_values(["data", "periodo", "setor", "horario", "nome"]).reset_index(drop=True), reasons


def suggest_extras(
    schedule: pd.DataFrame,
    colaboradores: pd.DataFrame,
    ideal: pd.DataFrame,
    start: date,
    domingo_especial: bool = False,
) -> pd.DataFrame:
    """Compatibilidade: usa a nova rotina de preenchimento automático."""
    filled, _ = auto_fill_coverage(schedule, colaboradores, ideal, pd.DataFrame(), pd.DataFrame(), start, domingo_especial=domingo_especial)
    return filled

def generate_schedule(
    colaboradores: pd.DataFrame,
    ideal: pd.DataFrame,
    eventos: pd.DataFrame,
    ajustes: pd.DataFrame,
    start: date,
    domingo_especial: bool = False,
    sugerir_extras: bool = True,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    base = generate_base_schedule(colaboradores, eventos, ajustes, start, domingo_especial=domingo_especial)
    reasons: dict[tuple[str, str, str], str] = {}
    if sugerir_extras:
        base, reasons = auto_fill_coverage(
            base,
            colaboradores,
            ideal,
            eventos,
            ajustes,
            start,
            domingo_especial=domingo_especial,
        )
    summary = coverage_summary(base, ideal, reasons=reasons)
    return base, summary


def whatsapp_text(schedule: pd.DataFrame, summary: pd.DataFrame | None = None) -> str:
    if schedule.empty:
        return "Nenhuma escala gerada."
    lines: list[str] = []
    for current, day_df in schedule.groupby("data", sort=True):
        dt = pd.to_datetime(current).date()
        dia = DIAS_PT[dt.weekday()].upper()
        lines.append(f"*{dia} — {dt.strftime('%d/%m/%Y')}*")
        for periodo in PERIODOS:
            per_df = day_df[day_df["periodo"] == periodo]
            if per_df.empty:
                continue
            lines.append(f"\n*{periodo.upper()}*")
            for setor, setor_df in per_df.groupby("setor", sort=False):
                for _, row in setor_df.sort_values(["horario", "nome"]).iterrows():
                    horario = str(row.get("horario", "")).strip()
                    nome = str(row.get("nome", "")).strip()
                    funcao = str(row.get("funcao", "")).strip()
                    origem = str(row.get("origem", "")).strip()
                    extra_tag = " (extra)" if origem == "Sugestão extra" else ""
                    funcao_txt = f" — {funcao}" if funcao and funcao.lower() not in {setor.lower(), "garçom"} else ""
                    lines.append(f"{setor}{funcao_txt} — {horario} — {nome}{extra_tag}")
        if summary is not None and not summary.empty:
            day_summary = summary[(summary["dia"].str.upper() == dia) & (summary["faltam"] > 0)]
            if not day_summary.empty:
                lines.append("\n*ALERTAS DE COBERTURA:*")
                for _, s in day_summary.iterrows():
                    motivo = str(s.get("motivo_falta", "")).strip() or "opções automáticas esgotadas"
                    lines.append(
                        f"Dia: {s['dia']} | Período: {s['periodo']} | Setor: {s['setor']} | "
                        f"Ideal: {int(s['ideal'])} | Escalado: {int(s['escalado'])} | "
                        f"Faltante: {int(s['faltam'])} | Motivo: {motivo}"
                    )
        lines.append("\n")
    return "\n".join(lines).strip()


def build_excel_sheets(
    schedule: pd.DataFrame,
    summary: pd.DataFrame,
    colaboradores: pd.DataFrame | None = None,
    eventos: pd.DataFrame | None = None,
    start: date | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    dias_semana = ["Sábado", "Domingo", "Segunda", "Terça", "Quarta", "Quinta", "Sexta"]
    datas_por_dia = {DIAS_PT[d.weekday()]: d for d in date_range(start)} if start else {}

    nomes = set(schedule["nome"].astype(str).str.strip()) if not schedule.empty and "nome" in schedule else set()
    if colaboradores is not None and not colaboradores.empty and "nome" in colaboradores:
        nomes.update(colaboradores["nome"].astype(str).str.strip().replace("", pd.NA).dropna().tolist())

    rows_colaborador = []
    for nome in sorted(n for n in nomes if n):
        emp_df = schedule[schedule["nome"].astype(str).str.strip().str.lower() == nome.lower()] if not schedule.empty else pd.DataFrame()
        row = {"Colaborador": nome}
        for dia in dias_semana:
            dia_df = emp_df[emp_df["dia"] == dia].sort_values(["horario", "setor", "funcao"]) if not emp_df.empty else pd.DataFrame()
            status = ""
            obs_status = ""
            if eventos is not None and dia in datas_por_dia:
                status, obs_status = employee_event_status(nome, datas_por_dia[dia], eventos)
            if status:
                row[dia] = "\n".join([part for part in [status, obs_status] if part])
            elif dia_df.empty:
                row[dia] = "FOLGA"
            else:
                partes = []
                for _, item in dia_df.iterrows():
                    setor_funcao = " / ".join(
                        str(item.get(col, "")).strip()
                        for col in ["setor", "funcao"]
                        if str(item.get(col, "")).strip()
                    )
                    detalhe = [str(item.get("horario", "")).strip(), setor_funcao]
                    if str(item.get("origem", "")).strip() == "Sugestão extra":
                        detalhe.append("EXTRA")
                    if str(item.get("observacao", "")).strip():
                        detalhe.append(str(item.get("observacao", "")).strip())
                    partes.append("\n".join([p for p in detalhe if p]))
                row[dia] = "\n\n".join(partes)
        rows_colaborador.append(row)
    por_colaborador = pd.DataFrame(rows_colaborador, columns=["Colaborador", *dias_semana])

    if schedule.empty:
        por_dia = pd.DataFrame(columns=["Dia", "Data", "Período", "Setor", "Horário", "Colaborador", "Função", "Observação"])
    else:
        por_dia = schedule.copy()
        por_dia["Data"] = por_dia["data"].map(lambda value: pd.to_datetime(value).strftime("%d/%m/%Y") if str(value).strip() else "")
        por_dia["Observação"] = por_dia.apply(
            lambda row: " | ".join(
                part for part in ["EXTRA" if str(row.get("origem", "")).strip() == "Sugestão extra" else "", str(row.get("observacao", "")).strip()] if part
            ),
            axis=1,
        )
        por_dia = por_dia.sort_values(["data", "periodo", "setor", "horario", "nome"])
        por_dia = por_dia.rename(columns={"dia": "Dia", "periodo": "Período", "setor": "Setor", "horario": "Horário", "nome": "Colaborador", "funcao": "Função"})
        por_dia = por_dia[["Dia", "Data", "Período", "Setor", "Horário", "Colaborador", "Função", "Observação"]]

    cobertura = summary.copy() if summary is not None else pd.DataFrame()
    if not cobertura.empty:
        cobertura["diferença"] = cobertura["escalado"].astype(int) - cobertura["ideal"].astype(int)
        cobertura["status"] = cobertura["diferença"].map(lambda diff: "OK" if diff == 0 else ("FALTA" if diff < 0 else "EXCEDENTE"))
        cobertura = cobertura.rename(columns={"dia": "Dia", "periodo": "Período", "setor": "Setor", "ideal": "Ideal", "escalado": "Escalado", "diferença": "Diferença", "status": "Status"})
        cobertura = cobertura[["Dia", "Período", "Setor", "Ideal", "Escalado", "Diferença", "Status"]]
    else:
        cobertura = pd.DataFrame(columns=["Dia", "Período", "Setor", "Ideal", "Escalado", "Diferença", "Status"])
    return por_colaborador, por_dia, cobertura


def _style_worksheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E3D")
    header_font = Font(color="FFFFFF", bold=True)
    day_fill = PatternFill("solid", fgColor="D9EAD3")
    folga_fill = PatternFill("solid", fgColor="FFF2CC")
    away_fill = PatternFill("solid", fgColor="DDEBF7")
    extra_fill = PatternFill("solid", fgColor="93C47D")
    alert_fill = PatternFill("solid", fgColor="F4CCCC")
    ok_fill = PatternFill("solid", fgColor="D9EAD3")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sheet.freeze_panes = "A2"
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            text = str(cell.value or "").upper()
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif "FOLGA" in text:
                cell.fill = folga_fill
            elif "FÉRIAS" in text or "FERIAS" in text or "AFASTAMENTO" in text:
                cell.fill = away_fill
            elif "EXTRA" in text:
                cell.fill = extra_fill
            elif "FALTA" in text:
                cell.fill = alert_fill
            elif text == "OK":
                cell.fill = ok_fill
            elif sheet.title == "Escala por Colaborador" and cell.column > 1:
                cell.fill = day_fill
    for column_cells in sheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_len + 2, 14), 45)
    for row in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row].height = 48


def to_excel_bytes(
    schedule: pd.DataFrame,
    summary: pd.DataFrame,
    colaboradores: pd.DataFrame | None = None,
    eventos: pd.DataFrame | None = None,
    start: date | None = None,
) -> bytes:
    output = io.BytesIO()
    por_colaborador, por_dia, cobertura = build_excel_sheets(schedule, summary, colaboradores, eventos, start)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        por_colaborador.to_excel(writer, index=False, sheet_name="Escala por Colaborador")
        por_dia.to_excel(writer, index=False, sheet_name="Escala por Dia")
        cobertura.to_excel(writer, index=False, sheet_name="Cobertura")
        for sheet in writer.book.worksheets:
            _style_worksheet(sheet)
    return output.getvalue()
